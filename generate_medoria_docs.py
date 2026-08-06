#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate Medoria architecture docs: Excel + RTF."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUT_XLSX = "/root/Medoria_Архитектура.xlsx"
OUT_RTF = "/root/Medoria_Архитектура.rtf"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_table(ws, headers, rows, start_row=1):
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row)
    for r, row in enumerate(rows, start_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(50, len(str(h)) + 8))


def sheet_overview(wb):
    ws = wb.active
    ws.title = "Обзор"
    ws["A1"] = "Архитектура Medoria — онбординг"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Дата: {date.today().strftime('%d.%m.%Y')}"
    ws.merge_cells("A1:F1")
    rows = [
        ("Компонент", "Адрес / хост", "Назначение"),
        ("Сервер приложений", "109.73.201.4 (server109)", "Docker, Caddy, PHP, LiveKit, gRPC, почта, Redmine"),
        ("Сервер БД", "46.149.69.33:5432", "PostgreSQL: medoria, medoria_test, redmine"),
        ("GitLab", "gitlab.com/medoria, gitlab.com/grpc9032312", "Исходники Dart, CI/CD, Container Registry"),
        ("GitHub", "— уточнить у команды", "Возможно мобильные клиенты / зеркала"),
        ("Основной портал", "connect.medoria.ru", "PHP Apache, API, LiveKit webhook"),
        ("Партнёры", "partner.medoria.ru", "PHP Apache"),
        ("Видео", "mediaserver.medoria.ru", "LiveKit WebRTC + TURN"),
        ("Realtime WS", "rt.medoria.ru/ws", "WebSocket зал ожидания (websocket_dart)"),
        (".NET API", "medoriawebapi.twc1.net", "medwebapi2 + локальный MSSQL на 109"),
        ("Задачи", "redmine.medoria.ru", "Redmine → PostgreSQL на 46.149"),
        ("Почта", "mail.connect.medoria.ru", "docker-mailserver"),
    ]
    write_table(ws, rows[0], rows[1:], start_row=4)
    ws["A18"] = "Важно: пароли и ключи не включены. Запросите у администратора."
    ws["A18"].font = Font(italic=True, color="C00000")


def sheet_servers(wb):
    ws = wb.create_sheet("Серверы")
    write_table(
        ws,
        ("Параметр", "109.73.201.4", "46.149.69.33"),
        [
            ("Роль", "Приложения, медиа, почта", "Центральная БД PostgreSQL"),
            ("SSH", "ssh server109 (алиас)", "Только с разрешения админа"),
            ("Hostname VPS", "4501149-is47636", "—"),
            ("RAM (ориентир)", "~11 GB", "—"),
            ("Диск (ориентир)", "~99 GB, следить за заполнением", "—"),
            ("Основные сервисы", "Caddy, site, livekit, grpc, mail", "PostgreSQL 5432"),
            ("Базы", "MSSQL локально (medwebapi2)", "medoria, medoria_test, redmine"),
            ("Пользователь БД", "—", "medoria (пароль у админа)"),
        ],
    )


def sheet_domains(wb):
    ws = wb.create_sheet("Домены")
    write_table(
        ws,
        ("Домен", "Бэкенд (Docker)", "Порт / путь", "Назначение"),
        [
            ("connect.medoria.ru", "site:80", "HTTPS", "Основной портал PHP"),
            ("partner.medoria.ru", "site:80", "HTTPS", "Партнёрский кабинет"),
            ("redmine.medoria.ru", "redmine:3000", "HTTPS", "Управление задачами"),
            ("medoriawebapi.twc1.net", "medwebapi2:8080", "HTTPS", ".NET REST API"),
            ("rt.medoria.ru", "server_dart:2065", "/ws", "WebSocket сигналинг"),
            ("mediaserver.medoria.ru", "172.20.0.1:7880, 7881", "/ws, /rtc", "LiveKit API + WebRTC"),
            ("mail.connect.medoria.ru", "mailserver", "25, 587, 993", "SMTP/IMAP"),
        ],
    )
    ws["A10"] = "Caddyfile: /var/www/Caddyfile"
    ws["A11"] = "SSL: Let's Encrypt (Caddy)"


def sheet_docker(wb):
    ws = wb.create_sheet("Docker")
    write_table(
        ws,
        ("Контейнер", "Образ / сборка", "Сеть", "Примечания"),
        [
            ("caddy", "caddy:2-alpine", "medoria_network", "Reverse proxy :80 :443"),
            ("site", "build connect.../apache", "medoria_network", "PHP Apache"),
            ("livekit", "livekit-server v1.9.7", "host", "node_ip 109.73.201.4, UDP 50000-53000"),
            ("livekitegress", "livekit/egress", "medoria_network", "Запись, API 127.0.0.1:9090"),
            ("redis", "redis:6-alpine", "medoria_network", "127.0.0.1:6379"),
            ("grpc_server", "registry.../server-grpc", "grpc + medoria", "Порты 50051-50053"),
            ("websocket_dart", "registry.../web-socket-dart", "grpc + medoria", "Алиас server_dart для Caddy"),
            ("cron_push_grpc_dart", "registry.../medoria/cron_push", "—", "/var/dart/www/cron_push_grpc_dart"),
            ("cron_status_grpc_dart", "registry.../cron_status", "—", "/var/dart/www/cron_status_grpc_dart"),
            ("mailserver", "docker-mailserver", "medoria_network", "Почта connect.medoria.ru"),
            ("www-redmine-1", "redmine:4.2", "medoria_network", "БД на 46.149.69.33"),
            ("medwebapi2", "build TestingDocker", "medoria_network", ".NET API"),
            ("mssql_container", "mssql/server", "medoria_network", "1435→1433, только для API"),
        ],
    )
    ws["A16"] = "Сеть: docker network create medoria_network (external)"


def sheet_paths(wb):
    ws = wb.create_sheet("Пути на сервере")
    write_table(
        ws,
        ("Путь", "Содержимое"),
        [
            ("/var/www/Caddyfile", "Маршрутизация всех доменов"),
            ("/var/www/docker-compose.yml", "Caddy, Redmine, mailserver"),
            ("/var/www/connect.medoria.ru/docker_php/", "PHP, site, server_dart исходники"),
            ("/var/www/livekit/", "livekit.yaml, docker-compose, egress"),
            ("/var/www/medoriawebapi.twc1.net/", ".NET API + MSSQL compose"),
            ("/var/dart/www/grpc_server/", "compose websocket_dart"),
            ("/var/dart/www/cron_push_grpc_dart/", "compose push cron"),
            ("/var/dart/www/cron_status_grpc_dart/", "compose status cron"),
            ("/var/www/Makefile", "Деплой по SSH (make restart-server)"),
        ],
    )


def sheet_gitlab(wb):
    ws = wb.create_sheet("GitLab")
    write_table(
        ws,
        ("Репозиторий / образ", "URL", "Деплой на сервер"),
        [
            ("cron_push_grpc_dart", "gitlab.com/medoria/cron_push_grpc_dart", "registry.gitlab.com/medoria/cron_push_grpc_dart:latest"),
            ("cron_status_grpc_dart", "gitlab.com/medoria/cron_status_grpc_dart", "registry.gitlab.com/medoria/cron_status_grpc_dart:latest"),
            ("server-grpc", "gitlab.com/grpc9032312/server-grpc", "registry.gitlab.com/grpc9032312/server-grpc:latest"),
            ("web-socket-dart", "gitlab.com/grpc9032312/web-socket-dart", "registry.gitlab.com/grpc9032312/web-socket-dart:latest"),
            ("PHP монолит", "Уточнить в GitLab (на сервере .git нет)", "rsync / CI → /var/www/connect.medoria.ru/docker_php/"),
        ],
    )
    ws["A8"] = "Обновление Dart: cd /var/dart/www/<проект> && docker compose pull && docker compose up -d"


def sheet_video_flow(wb):
    ws = wb.create_sheet("Видеозвонок")
    write_table(
        ws,
        ("Шаг", "Сервис", "Описание"),
        [
            ("1", "connect.medoria.ru", "PHP создаёт комнату/токен LiveKit (livekit/)"),
            ("2", "rt.medoria.ru/ws", "Пациент в зале ожидания (WebSocket)"),
            ("3", "mediaserver.medoria.ru", "WebRTC: UDP 50000-53000, TURN 3478/5349"),
            ("4", "livekitegress", "Запись консультации по API"),
            ("5", "connect.medoria.ru/livekit/", "Webhook от LiveKit"),
        ],
    )
    ws["A8"] = "Конфиг LiveKit: /var/www/livekit/livekit.yaml (rtc.node_ip: 109.73.201.4)"


def sheet_diagnostics(wb):
    ws = wb.create_sheet("Диагностика")
    write_table(
        ws,
        ("Симптом", "Команда / лог"),
        [
            ("Список контейнеров", "docker ps"),
            ("Сайт 502", "docker logs caddy; docker logs site"),
            ("WS зал ожидания", "docker logs websocket_dart; grep rt.medoria в логах caddy"),
            ("Обрыв видео", "docker logs livekit; docker logs livekitegress"),
            ("Почта", "docker logs mailserver"),
            ("Память", "free -h; pgrep -c docker-proxy"),
            ("Перезагрузка Caddy", "docker exec caddy caddy reload --config /etc/caddy/Caddyfile"),
            ("Перезагрузка LiveKit", "cd /var/www/livekit && docker compose restart livekit"),
        ],
    )


def sheet_checklist(wb):
    ws = wb.create_sheet("Чеклист")
    write_table(
        ws,
        ("№", "Действие", "Статус"),
        [
            ("1", "Получить SSH (server109) и ключи", ""),
            ("2", "Доступ GitLab: medoria, grpc9032312", ""),
            ("3", "Доступ PostgreSQL 46.149.69.33 (read-only для старта)", ""),
            ("4", "Прочитать /var/www/Caddyfile на сервере", ""),
            ("5", "Различать PostgreSQL (46.149) и MSSQL (109)", ""),
            ("6", "Локально поднять docker_php или Dart-репо", ""),
            ("7", "Перед prod: бэкап конфигов (*.bak-YYYYMMDD)", ""),
        ],
    )


def rtf_escape(s):
    return s.replace("\\", "\\\\").replace("{", "\\{").replace("}", "\\}")


def build_rtf():
    sections = [
        ("Архитектура Medoria — онбординг", f"Дата: {date.today().strftime('%d.%m.%Y')}\\par\\par"),
        ("1. Общая схема", r"""
Платформа Medoria — телемедицина: веб-порталы, мобильные клиенты (Flutter), видеоконсультации, протоколы, почта, Redmine.

Сервер приложений: 109.73.201.4 (SSH: server109)
Сервер БД: 46.149.69.33 (PostgreSQL, порт 5432)
GitLab: gitlab.com/medoria и gitlab.com/grpc9032312
GitHub: уточнить у команды (на prod не используется напрямую)
\par"""),
        ("2. Домены", r"""
connect.medoria.ru — основной PHP-портал
partner.medoria.ru — партнёрский кабинет
rt.medoria.ru/ws — WebSocket (зал ожидания)
mediaserver.medoria.ru — LiveKit (видео)
medoriawebapi.twc1.net — .NET API
redmine.medoria.ru — задачи
mail.connect.medoria.ru — почта
\par"""),
        ("3. Видеозвонок", r"""
1. PHP на connect создаёт токен LiveKit
2. Пациент: WebSocket rt.medoria.ru
3. Медиа: mediaserver.medoria.ru (WebRTC UDP 50000-53000)
4. Запись: livekitegress
Конфиг: /var/www/livekit/livekit.yaml, node_ip 109.73.201.4
\par"""),
        ("4. GitLab Registry", r"""
registry.gitlab.com/medoria/cron_push_grpc_dart:latest
registry.gitlab.com/medoria/cron_status_grpc_dart:latest
registry.gitlab.com/grpc9032312/server-grpc:latest
registry.gitlab.com/grpc9032312/web-socket-dart:latest

Обновление: docker compose pull && docker compose up -d в /var/dart/www/...
\par"""),
        ("5. Базы данных", r"""
PostgreSQL 46.149.69.33 — основные данные (medoria, medoria_test, redmine)
MSSQL на 109.73.201.4 — только для medwebapi2 (порт 1435)
\par"""),
        ("6. Безопасность", r"""
Пароли в этот документ не включены. Не публикуйте /var/www/Makefile с учётными данными.
\par"""),
    ]
    body = []
    for title, text in sections:
        body.append(r"{\b\fs28 " + rtf_escape(title) + r"}\par")
        body.append(text)
    content = "".join(body)
    return (
        r"{\rtf1\ansi\deff0{\fonttbl{\f0 Arial;}}"
        r"\f0\fs22 "
        + content
        + "}"
    )


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_servers(wb)
    sheet_domains(wb)
    sheet_docker(wb)
    sheet_paths(wb)
    sheet_gitlab(wb)
    sheet_video_flow(wb)
    sheet_diagnostics(wb)
    sheet_checklist(wb)
    wb.save(OUT_XLSX)

    with open(OUT_RTF, "w", encoding="utf-8") as f:
        f.write(build_rtf())

    print(f"Created: {OUT_XLSX}")
    print(f"Created: {OUT_RTF}")


if __name__ == "__main__":
    main()
